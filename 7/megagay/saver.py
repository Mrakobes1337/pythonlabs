import openpyxl as op

from .inf import odezhdainf
try:
    wb=op.open("test.xlsx",data_only=True)
    ws=wb.active
    first=False
except:
    wb=op.Workbook()
    ws=wb.active
    first=True
    
info={"name":1,"quan":2}
i=3
for nm in odezhdainf:
    
    od=odezhdainf[nm]
    for furtype in od["types"]:
        strn=f'{od["name"]}_{furtype}'
        if first:
            ws.cell(row=1,column=i).value=strn
        info[strn]=i
        i+=1
    
if first:
    ws["A1"]="Тип одежды"
    ws["B1"]="Количество"
    ws.cell(row=1,column=i).value="Стоимость"
info["price"]=i
    

def editor(tosave):
    ret={"1":tosave["type"],
         "2":tosave["quan"],
         str(i):tosave["price"]}
    j=3
    for el in tosave:
        if el in odezhdainf:
            odname=odezhdainf[el]["name"]
            odtype=tosave[f"{el}_type"]
            z=str(info[f"{odname}_{odtype}"])
            ret[z]=tosave[el]
    return ret

def save(tosave):
    norm=editor(tosave)
    j=0
    zxc=1
    while zxc != None:
        j+=1
        zxc=ws[f"A{j}"].value
        
    for l in range(1,i+1):
        if str(l) not in norm:
            ws.cell(row=j,column=l).value=0
        else:
            ws.cell(row=j,column=l).value=norm[str(l)]
    wb.save("test.xlsx")


